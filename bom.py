import creopyson
import re
import openpyxl
import os
c = creopyson.Client()
c.connect()

c.creo_cd("C:\\Users\\patrick.vieira\\Documents\\Creo\\TKZ")

file = "asm0001.asm"

x = creopyson.bom.get_paths(c, file_=file, paths=None, skeletons=None, top_level=True, get_transforms=None, exclude_inactive=None)

#print(x)

#regex para paineis

from openpyxl import load_workbook
import datetime
wb = load_workbook('TESTE_PAINEL.xlsx')
sheet = wb.active
j = 2
painel = 'EF380789C0'
#sheet.cell(row=1, column=1).value = 'CÓDIGO'
#sheet.cell(row=1, column=2).value = 'QUANTIDADE'
date = datetime.datetime.strptime('03/10/2019', "%m/%d/%Y")

for i in range(18):
	cont_painel = painel + str(i).zfill(2)
	y = re.findall(cont_painel, str(x))
	if len(y) != 0:
		#print(cont_painel + '-' + str(len(y)))
		sheet.cell(row=j, column=4).value = (j-1)*10
		sheet.cell(row=j, column=5).value = 1
		sheet.cell(row=j, column=7).value = cont_painel
		sheet.cell(row=j, column=8).value = len(y)
		sheet.cell(row=j, column=9).value = 'pcs'
		sheet.cell(row=j, column=10).value = date
		j = j + 1
	else:
		pass

wb.save('TESTE_PAINEL.xlsx')
os.startfile("C:\\Users\\patrick.vieira\\Documents\\TESTE_PAINEL.xlsx")
